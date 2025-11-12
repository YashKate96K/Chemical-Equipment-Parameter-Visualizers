import csv
import io
from collections import Counter
from statistics import mean, median
import numpy as np
from sklearn.cluster import KMeans
from math import sqrt
from openpyxl import load_workbook

REQUIRED_COLUMNS = ['Equipment Name', 'Type', 'Flowrate', 'Pressure', 'Temperature']

class CSVValidationError(Exception):
    pass

def _read_tabular(file_bytes: bytes, filename: str):
    """Return (header, rows) from CSV or XLSX file. Rows are list[dict]."""
    name = (filename or '').lower()
    if name.endswith('.xlsx'):
        bio = io.BytesIO(file_bytes)
        wb = load_workbook(bio, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
        except StopIteration:
            return [], []
        header = [h for h in header if h != '']
        rows = []
        for ridx, vals in enumerate(rows_iter, start=1):
            row = {header[i]: vals[i] for i in range(min(len(header), len(vals)))}
            rows.append(row)
        return header, rows
    else:
        text = file_bytes.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        header = reader.fieldnames or []
        rows = list(reader)
        return header, rows


def parse_and_validate(file_bytes: bytes, filename: str | None = None):
    header, rows = _read_tabular(file_bytes, filename or '')
    if not header or any(col not in header for col in REQUIRED_COLUMNS):
        raise CSVValidationError('File must include columns: ' + ', '.join(REQUIRED_COLUMNS))

    # Validate that base numeric columns are parseable when present (allow missing)
    for idx, row in enumerate(rows, start=1):
        for base in ('Flowrate', 'Pressure', 'Temperature'):
            val = row.get(base)
            if val in (None, ''):
                continue
            try:
                float(val)
            except Exception:
                raise CSVValidationError(f'Invalid numeric in {base} at row {idx}')

    # Type distribution
    type_counts = Counter()
    for row in rows:
        type_counts[str(row.get('Type', 'Unknown'))] += 1

    # Numeric column detection allowing missing values: consider numeric if >=80% of non-empty values are parseable
    numeric_cols = []
    for col in header:
        if col in ('Type', 'Equipment Name'):
            continue
        vals = [row.get(col) for row in rows]
        non_empty = [v for v in vals if v not in (None, '')]
        if not non_empty:
            continue
        parsable = 0
        parsed_vals = []
        for v in non_empty:
            try:
                fv = float(v)
                parsable += 1
                parsed_vals.append(fv)
            except Exception:
                pass
        if parsable / len(non_empty) >= 0.8 and parsed_vals:
            numeric_cols.append((col, parsed_vals))

    if not numeric_cols:
        # fallback to core three if present
        for base_col in ['Flowrate', 'Pressure', 'Temperature']:
            try:
                parsed = [float(r[base_col]) for r in rows if r.get(base_col) not in (None, '')]
                if parsed:
                    numeric_cols.append((base_col, parsed))
            except Exception:
                continue

    averages = {name: mean(vals) for name, vals in numeric_cols}
    medians = {name: median(vals) for name, vals in numeric_cols}
    mins = {name: min(vals) for name, vals in numeric_cols}
    maxs = {name: max(vals) for name, vals in numeric_cols}

    summary = {
        'averages': averages,
        'median': medians,
        'min': mins,
        'max': maxs,
        'type_distribution': dict(type_counts),
        'row_count': len(rows),
        'numeric_columns': list(averages.keys()),
        'all_columns': header,
    }

    # Build preview csv from header/rows (first 10)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=header)
    writer.writeheader()
    for row in rows[:10]:
        writer.writerow({k: row.get(k, '') for k in header})
    preview_csv = output.getvalue()

    return summary, preview_csv


def parse_rows(file_bytes: bytes, filename: str | None = None):
    """Return list of rows with numeric fields converted to float and preserve other columns.
    Adds a 1-based 'Record' index for display.
    """
    header, raw_rows = _read_tabular(file_bytes, filename or '')
    rows = []
    for i, row in enumerate(raw_rows, start=1):
        # Convert known numeric fields if present
        for col in header:
            if col in ('Flowrate', 'Pressure', 'Temperature'):
                try:
                    v = row.get(col)
                    row[col] = float(v) if v not in (None, '') else None
                except Exception:
                    row[col] = None
        row['Record'] = i
        rows.append(row)
    return header, rows


def detect_anomalies_zscore(rows, numeric_cols=('Flowrate', 'Pressure', 'Temperature')):
    """Simple anomaly detection: mark records with any z-score > 3 on specified numeric columns.
    Returns set of 1-based Record indices considered anomalies and per-column stats.
    """
    stats = {}
    anomalies = set()
    for col in numeric_cols:
        vals = [r[col] for r in rows if isinstance(r.get(col), (int, float))]
        if not vals:
            continue
        mu = mean(vals)
        # std (population) safe guard
        var = mean([(v - mu) ** 2 for v in vals]) if len(vals) > 1 else 0.0
        std = sqrt(var)
        stats[col] = {'mean': mu, 'std': std}
        if std == 0:
            continue
        for r in rows:
            v = r.get(col)
            if isinstance(v, (int, float)):
                z = abs((v - mu) / std) if std else 0
                if z > 3:
                    anomalies.add(r['Record'])
    return anomalies, stats


def infer_column_types(header, rows):
    types = {}
    for col in header:
        if col == 'Record':
            continue
        numeric = True
        for r in rows:
            v = r.get(col)
            if v in (None, ''):
                continue
            try:
                float(v)
            except Exception:
                numeric = False
                break
        types[col] = 'numeric' if numeric else 'string'
    return types


def compute_quality(header, rows, previous_headers=None):
    # Missing values per column
    missing = {col: 0 for col in header}
    for r in rows:
        for col in header:
            if r.get(col) in (None, ''):
                missing[col] += 1

    # Duplicate rows (stringify rows for simple detection)
    seen = set()
    dup_count = 0
    samples = []
    for r in rows:
        key = tuple((col, str(r.get(col))) for col in header)
        if key in seen:
            dup_count += 1
            if len(samples) < 3:
                samples.append({c: r.get(c) for c in header})
        else:
            seen.add(key)

    # Ranges for numeric columns
    ranges = {}
    for col in header:
        try:
            vals = [float(r.get(col)) for r in rows if r.get(col) not in (None, '')]
        except Exception:
            vals = []
        if vals:
            ranges[col] = {'min': min(vals), 'max': max(vals)}

    # Schema drift vs previous headers
    drift = {}
    if previous_headers:
        prev = set(previous_headers)
        cur = set(header)
        drift = {
            'added_columns': sorted(list(cur - prev)),
            'removed_columns': sorted(list(prev - cur)),
            'unchanged_columns': sorted(list(cur & prev)),
        }

    return {
        'missing_values': missing,
        'duplicate_rows': {'count': dup_count, 'samples': samples},
        'ranges': ranges,
        'column_types': infer_column_types(header, rows),
        'schema_drift': drift,
    }


def compute_correlations(rows, numeric_cols):
    data = []
    for r in rows:
        try:
            data.append([float(r.get(c)) if r.get(c) not in (None, '') else np.nan for c in numeric_cols])
        except Exception:
            data.append([np.nan for _ in numeric_cols])
    arr = np.array(data, dtype=float)
    # drop rows with any nan
    arr = arr[~np.isnan(arr).any(axis=1)]
    if arr.shape[0] < 2:
        return {'matrix': [], 'order': numeric_cols, 'strongest_pairs': []}
    corr = np.corrcoef(arr, rowvar=False)
    # strongest pairs
    pairs = []
    n = len(numeric_cols)
    for i in range(n):
        for j in range(i+1, n):
            pairs.append(((numeric_cols[i], numeric_cols[j]), float(corr[i,j])))
    pairs.sort(key=lambda x: abs(x[1]), reverse=True)
    return {
        'matrix': corr.tolist(),
        'order': numeric_cols,
        'strongest_pairs': [{'cols': p[0], 'corr': p[1]} for p in pairs[:3]],
    }


def compute_variance_skewness(rows, numeric_cols):
    stats = {}
    for c in numeric_cols:
        vals = [float(r.get(c)) for r in rows if r.get(c) not in (None, '')]
        if len(vals) >= 2:
            v = np.var(vals)
            mean_v = np.mean(vals)
            std = np.std(vals) or 1e-9
            # Fisher-Pearson sample skewness approximation
            skew = float(np.mean(((np.array(vals)-mean_v)/std)**3))
            stats[c] = {'variance': float(v), 'skewness': skew}
        elif vals:
            stats[c] = {'variance': 0.0, 'skewness': 0.0}
    return stats


def kmeans_clusters(rows, numeric_cols, max_k=4):
    # build matrix
    X = []
    rec_ids = []
    for r in rows:
        try:
            row_vals = [float(r.get(c)) for c in numeric_cols]
            if any(v in (None, '') for v in row_vals):
                continue
            X.append(row_vals)
            rec_ids.append(r['Record'])
        except Exception:
            continue
    if len(X) < 2:
        return {'k': 0, 'labels': [], 'rec_ids': [], 'centroids': []}
    X = np.array(X, dtype=float)
    # choose k with simple elbow heuristic over 2..max_k
    best_k, best_inertia, best_labels, best_centroids = 2, None, None, None
    for k in range(2, min(max_k, len(X)) + 1):
        km = KMeans(n_clusters=k, n_init=10, random_state=42)
        km.fit(X)
        inertia = km.inertia_
        if best_inertia is None or inertia < best_inertia:
            best_k, best_inertia = k, inertia
            best_labels, best_centroids = km.labels_.tolist(), km.cluster_centers_.tolist()
    return {'k': best_k, 'labels': best_labels, 'rec_ids': rec_ids, 'centroids': best_centroids}
